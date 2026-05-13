from __future__ import annotations

import datetime as dt
import uuid
from io import BytesIO
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook

from apps.hospital_directory.models import Department, Designation, WorkingSite
from apps.imports.models import ImportBatch
from apps.users.services import provisioning
from core.constants import Gender, ImportBatchStatus, ImportBatchType


def _parse_date(value: Any):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return dt.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Unrecognized date value: {value!r}")


def _normalize_header(row: tuple[Any, ...]) -> list[str]:
    return [str(c).strip().lower().replace(" ", "_") if c is not None else "" for c in row]


def _resolve_staff_directory_row(
    row: tuple[Any, ...],
    idx: dict[str, int],
) -> tuple[Department, Designation, WorkingSite]:
    def col(name: str, required: bool = True) -> int:
        if name not in idx:
            if required:
                raise ValueError(f"Missing column {name!r} in spreadsheet.")
            return -1
        return idx[name]

    c_did = col("department_id", required=False)
    c_desid = col("designation_id", required=False)
    c_sid = col("working_site_id", required=False)
    c_dept = col("department", required=False)
    c_des = col("designation", required=False)
    c_site = col("working_site", required=False)

    def cell(ci: int) -> str:
        if ci < 0 or ci >= len(row) or row[ci] is None:
            return ""
        return str(row[ci]).strip()

    if c_did >= 0 and cell(c_did):
        dept = Department.objects.get(pk=uuid.UUID(cell(c_did)))
    elif c_dept >= 0 and cell(c_dept):
        dept = Department.objects.filter(name__iexact=cell(c_dept)).first()
        if dept is None:
            raise ValueError(f"Unknown department: {cell(c_dept)!r}")
    else:
        raise ValueError("Provide department_id or department.")

    if c_desid >= 0 and cell(c_desid):
        des = Designation.objects.get(pk=uuid.UUID(cell(c_desid)))
    elif c_des >= 0 and cell(c_des):
        des = Designation.objects.filter(name__iexact=cell(c_des)).first()
        if des is None:
            raise ValueError(f"Unknown designation: {cell(c_des)!r}")
    else:
        raise ValueError("Provide designation_id or designation.")

    if c_sid >= 0 and cell(c_sid):
        site = WorkingSite.objects.get(pk=uuid.UUID(cell(c_sid)))
    elif c_site >= 0 and cell(c_site):
        site = WorkingSite.objects.filter(name__iexact=cell(c_site)).first()
        if site is None:
            raise ValueError(f"Unknown working site: {cell(c_site)!r}")
    else:
        raise ValueError("Provide working_site_id or working_site.")

    return dept, des, site


@transaction.atomic
def import_hospital_staff_from_xlsx(
    *,
    file_bytes: bytes,
    file_name: str,
    acting_user,
) -> ImportBatch:
    """
    Expected headers (row 1):
    staff_number (or legacy employee_number), national_id, full_name, phone,
    date_employed, and either (department_id, designation_id, working_site_id) or
    (department, designation, working_site) names matching directory rows.
    hod_staff_number (optional; legacy hod_employee_number accepted)
    """
    batch = ImportBatch.objects.create(
        imported_by=acting_user,
        batch_type=ImportBatchType.HOSPITAL_STAFF,
        file_name=file_name,
        status=ImportBatchStatus.PROCESSING,
    )
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        batch.status = ImportBatchStatus.FAILED
        batch.save(update_fields=["status"])
        return batch
    headers = _normalize_header(rows[0])
    idx = {h: i for i, h in enumerate(headers) if h}

    def col(name: str, required: bool = True) -> int:
        if name not in idx:
            if required:
                raise ValueError(f"Missing column {name!r} in spreadsheet.")
            return -1
        return idx[name]

    def col_first(names: tuple[str, ...], *, required: bool = True) -> int:
        for name in names:
            if name in idx:
                return idx[name]
        if required:
            raise ValueError(
                f"Missing one of columns {list(names)!r} in spreadsheet."
            )
        return -1

    c_emp = col_first(("staff_number", "employee_number"))
    c_nid = col("national_id")
    c_name = col("full_name")
    c_phone = col("phone")
    c_date = col("date_employed")
    c_hod = col_first(
        ("hod_staff_number", "hod_employee_number"),
        required=False,
    )

    success = 0
    failed = 0
    data_rows = rows[1:]
    batch.total_rows = len(data_rows)
    batch.save(update_fields=["total_rows"])

    for row in data_rows:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            dept, des, site = _resolve_staff_directory_row(row, idx)
            provisioning.create_hospital_staff_user(
                acting_user=acting_user,
                staff_number=str(row[c_emp]).strip(),
                national_id=str(row[c_nid]).strip(),
                full_name=str(row[c_name]).strip(),
                department_id=dept.pk,
                designation_id=des.pk,
                working_site_id=site.pk,
                staff_role_id=None,
                phone=str(row[c_phone]).strip(),
                date_employed=_parse_date(row[c_date]),
            )
            success += 1
        except Exception:
            failed += 1

    batch.success_rows = success
    batch.failed_rows = failed
    batch.status = ImportBatchStatus.COMPLETED
    batch.save(update_fields=["success_rows", "failed_rows", "status"])
    return batch


@transaction.atomic
def import_students_from_xlsx(
    *,
    file_bytes: bytes,
    file_name: str,
    acting_user,
) -> ImportBatch:
    """
    Expected headers:
    registration_no, full_name, programme, faculty, year_of_study, phone, dob,
    university (optional), supervisor_registration_no (optional),
    contact_email (optional), gender (optional), hospital_department_id (optional),
    dashboard_notes (optional)
    """
    batch = ImportBatch.objects.create(
        imported_by=acting_user,
        batch_type=ImportBatchType.STUDENT,
        file_name=file_name,
        status=ImportBatchStatus.PROCESSING,
    )
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        batch.status = ImportBatchStatus.FAILED
        batch.save(update_fields=["status"])
        return batch
    headers = _normalize_header(rows[0])
    idx = {h: i for i, h in enumerate(headers) if h}

    def col(name: str, required: bool = True) -> int:
        if name not in idx:
            if required:
                raise ValueError(f"Missing column {name!r} in spreadsheet.")
            return -1
        return idx[name]

    c_reg = col("registration_no")
    c_name = col("full_name")
    c_prog = col("programme")
    c_fac = col("faculty")
    c_year = col("year_of_study")
    c_phone = col("phone")
    c_dob = col("dob")
    c_uni = col("university", required=False)
    c_sup = col("supervisor_registration_no", required=False)
    c_email = col("email", required=False)
    c_cemail = col("contact_email", required=False)
    c_gender = col("gender", required=False)
    c_hdept = col("hospital_department_id", required=False)
    c_notes = col("dashboard_notes", required=False)

    success = 0
    failed = 0
    data_rows = rows[1:]
    batch.total_rows = len(data_rows)
    batch.save(update_fields=["total_rows"])

    for row in data_rows:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            sup_user = None
            if c_sup >= 0 and c_sup < len(row) and row[c_sup]:
                sup_reg = str(row[c_sup]).strip()
                sup_user = get_user_model().objects.filter(username=sup_reg).first()
            uni = "Zanzibar University"
            if c_uni >= 0 and c_uni < len(row) and row[c_uni]:
                uni = str(row[c_uni]).strip()
            year_val = row[c_year]
            year_int = int(year_val) if year_val is not None else 0
            dob_raw = row[c_dob]
            dob_str = str(dob_raw).strip() if dob_raw is not None else ""
            if len(dob_str) == 10 and dob_str.count("/") == 2:
                d, m, y = dob_str.split("/")
                dob_str = f"{d.zfill(2)}{m.zfill(2)}{y}"
            elif len(dob_str) == 10 and dob_str.count("-") == 2:
                y, m, d = dob_str.split("-")
                dob_str = f"{d.zfill(2)}{m.zfill(2)}{y}"
            login_email = ""
            if c_email >= 0 and c_email < len(row) and row[c_email]:
                login_email = str(row[c_email]).strip()
            contact_email = ""
            if c_cemail >= 0 and c_cemail < len(row) and row[c_cemail]:
                contact_email = str(row[c_cemail]).strip()
            gender = Gender.UNSPECIFIED
            if c_gender >= 0 and c_gender < len(row) and row[c_gender]:
                g = str(row[c_gender]).strip().lower()
                allowed = {c.value for c in Gender}
                if g in allowed:
                    gender = g
            hosp_dept_id = None
            if c_hdept >= 0 and c_hdept < len(row) and row[c_hdept]:
                hosp_dept_id = uuid.UUID(str(row[c_hdept]).strip())
            notes = ""
            if c_notes >= 0 and c_notes < len(row) and row[c_notes]:
                notes = str(row[c_notes]).strip()
            provisioning.create_student_user(
                acting_user=acting_user,
                registration_no=str(row[c_reg]).strip(),
                full_name=str(row[c_name]).strip(),
                programme=str(row[c_prog]).strip(),
                faculty=str(row[c_fac]).strip(),
                year_of_study=year_int,
                phone=str(row[c_phone]).strip(),
                dob=dob_str,
                university=uni,
                supervisor_user=sup_user,
                email=login_email,
                contact_email=contact_email,
                gender=gender,
                hospital_department_id=hosp_dept_id,
                dashboard_notes=notes,
            )
            success += 1
        except Exception:
            failed += 1

    batch.success_rows = success
    batch.failed_rows = failed
    batch.status = ImportBatchStatus.COMPLETED
    batch.save(update_fields=["success_rows", "failed_rows", "status"])
    return batch
